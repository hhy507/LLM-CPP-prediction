#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Direct LLM CPP/non-CPP classification baseline (Reviewer 5, Comment 7).

Each peptide in the held-out test set is given directly to a current LLM, which
is asked to return a single digit: 1 = CPP, 0 = non-CPP. The accuracy is then
compared with the authors' ML models on the SAME 185-sequence test split.

API keys are read ONLY from environment variables (never hard-coded):
    DEEPSEEK_API_KEY      - DeepSeek official API
    OPENROUTER_API_KEY    - OpenRouter API
Optional model overrides:
    DEEPSEEK_MODEL    (default: deepseek-chat)
    OPENROUTER_MODEL  (default: openai/gpt-oss-20b)

Usage:
    python llm_direct_classification.py [--limit N] [--workers K]
"""
import os, re, csv, sys, json, time, math, argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import openpyxl

TEST_XLSX = "deno/RF/seq/data/test.xlsx"   # columns: ID, Label, Sequence (185 rows)

PROMPT = (
    "You are an expert in peptide biology and cell-penetrating peptides (CPPs). "
    "Decide whether the following peptide sequence is a cell-penetrating peptide. "
    "Answer with a single digit and nothing else: 1 if it is a CPP, 0 if it is not. "
    "Sequence: {seq}"
)

PROVIDERS = {
    "DeepSeek-V3": dict(
        url="https://api.deepseek.com/chat/completions",
        key_env="DEEPSEEK_API_KEY",
        model=os.environ.get("DEEPSEEK_MODEL", "deepseek-chat"),
        max_tokens=10,
    ),
    "GPT-4o": dict(
        url="https://openrouter.ai/api/v1/chat/completions",
        key_env="OPENROUTER_API_KEY",
        model=os.environ.get("OPENROUTER_MODEL", "openai/gpt-4o"),
        max_tokens=2000,   # large enough for any model (gpt-4o emits a single digit)
    ),
}


def parse_pred(text):
    """Map a model reply to 0/1; None if unparseable."""
    if not text:
        return None
    s = text.strip()
    if s and s[0] in "01":
        return int(s[0])
    digits = re.findall(r"[01]", s)
    if digits:
        return int(digits[-1])          # take the final conclusion digit
    low = s.lower()
    if "non" in low and "cpp" in low:
        return 0
    if "cpp" in low:
        return 1
    return None


def call(provider, seq, retries=4):
    p = PROVIDERS[provider]
    key = os.environ[p["key_env"]]
    body = {
        "model": p["model"],
        "messages": [{"role": "user", "content": PROMPT.format(seq=seq)}],
        "temperature": 0,
        "max_tokens": p["max_tokens"],
    }
    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    for attempt in range(retries):
        try:
            r = requests.post(p["url"], headers=headers, json=body, timeout=120)
            if r.status_code != 200:
                raise RuntimeError(f"HTTP {r.status_code}: {r.text[:120]}")
            d = r.json()
            content = d["choices"][0]["message"].get("content") or ""
            return parse_pred(content), content.strip()
        except Exception as e:
            if attempt == retries - 1:
                return None, f"ERROR:{e}"
            time.sleep(2 * (attempt + 1))


def metrics(rows, pred_key):
    pairs = [(r["label"], r[pred_key]) for r in rows if r[pred_key] is not None]
    n = len(pairs)
    abstain = len(rows) - n
    tp = sum(1 for t, p in pairs if t == 1 and p == 1)
    tn = sum(1 for t, p in pairs if t == 0 and p == 0)
    fp = sum(1 for t, p in pairs if t == 0 and p == 1)
    fn = sum(1 for t, p in pairs if t == 1 and p == 0)
    acc = (tp + tn) / n if n else 0.0
    prec = tp / (tp + fp) if (tp + fp) else 0.0
    rec = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0
    den = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = (tp * tn - fp * fn) / den if den else 0.0
    return dict(n=n, abstain=abstain, tp=tp, tn=tn, fp=fp, fn=fn,
                acc=acc, precision=prec, recall=rec, f1=f1, mcc=mcc)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="only first N sequences (0 = all)")
    ap.add_argument("--workers", type=int, default=6)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(TEST_XLSX, read_only=True)
    ws = wb.active
    head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    i_id, i_lab, i_seq = head.index("ID"), head.index("Label"), head.index("Sequence")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[i_seq] is None:
            continue
        rows.append({"id": r[i_id], "label": int(r[i_lab]), "seq": str(r[i_seq])})
    wb.close()
    if args.limit:
        rows = rows[: args.limit]
    print(f"Loaded {len(rows)} sequences from {TEST_XLSX}")

    for name in PROVIDERS:
        for r in rows:
            r[name] = None
            r[name + "_raw"] = ""

    tasks = [(name, idx) for name in PROVIDERS for idx in range(len(rows))]
    done = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        fut = {ex.submit(call, name, rows[idx]["seq"]): (name, idx) for name, idx in tasks}
        for f in as_completed(fut):
            name, idx = fut[f]
            pred, raw = f.result()
            rows[idx][name] = pred
            rows[idx][name + "_raw"] = raw
            done += 1
            if done % 20 == 0 or done == len(tasks):
                print(f"  {done}/{len(tasks)} calls done")

    # save per-sequence CSV
    out_csv = "llm_direct_classification_results.csv"
    cols = ["id", "label", "seq"] + sum([[n, n + "_raw"] for n in PROVIDERS], [])
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in cols})
    print(f"\nSaved per-sequence predictions -> {out_csv}")

    # metrics + comparison
    summary = {name: metrics(rows, name) for name in PROVIDERS}
    ml_baseline = {
        "RF-Fre":     dict(acc=0.946, precision=None, recall=None, f1=0.948, mcc=0.893),
        "RF-GPT-Fre": dict(acc=0.903, precision=None, recall=None, f1=0.906, mcc=0.806),
        "RF-DS-Fre":  dict(acc=0.897, precision=None, recall=None, f1=0.898, mcc=0.799),
    }
    report = {"test_set": TEST_XLSX, "n_total": len(rows),
              "models": {p: PROVIDERS[p]["model"] for p in PROVIDERS},
              "llm_direct": summary, "ml_baseline_same_split": ml_baseline}
    with open("llm_direct_classification_summary.json", "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, ensure_ascii=False)

    print("\n================ Direct-LLM classification (held-out 185) ================")
    print(f"{'Model':<28}{'ACC':>7}{'Prec':>7}{'Rec':>7}{'F1':>7}{'MCC':>7}{'abst':>6}")
    for name, m in summary.items():
        mdl = PROVIDERS[name]["model"]
        print(f"{name+' ('+mdl+')':<28}{m['acc']:>7.3f}{m['precision']:>7.3f}{m['recall']:>7.3f}{m['f1']:>7.3f}{m['mcc']:>7.3f}{m['abstain']:>6}")
    print("---- Authors' ML models (same 185 split) ----")
    for name, m in ml_baseline.items():
        print(f"{name:<28}{m['acc']:>7.3f}{'-':>7}{'-':>7}{m['f1']:>7.3f}{m['mcc']:>7.3f}{0:>6}")
    print("\nSaved summary -> llm_direct_classification_summary.json")


if __name__ == "__main__":
    main()
